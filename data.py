import mysql.connector
from SQL_handle import LogData,Get_DATA,Get_DATA_Reset,reset,insert_coconut_count,store_daily_resets,store_hourly_resets,store_monthly_resets,store_weekly_resets
import datetime , timedelta
import csv
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import schedule
import os
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
ct = datetime.now()

prev_month= ct.month

mydb = mysql.connector.connect(
        host = "localhost",
        user = "root",
        passwd = "Hashini@123",
        database = "coconut"
)

def GetDATA(cam_id,time,count):
    if cam_id ==1 :
        Get_DATA(time,count)

    elif cam_id ==2:
        Get_DATA_Reset(time,count)


    


# def LogData(time,count,days):
def Log_Data(cam_id, time,count):
     LogData(cam_id, time,count)





     
def Extract_hourly():
    count1=[]
    datet1=[]
    mycursor = mydb.cursor()
    query = "SELECT Datetime,Hour_Count FROM log_hourly ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    # time = result[0]
    # time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
    # result1= result[1]

    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet1.append(time)
    for result in results:
        count1.append(result[1])
    # print("RESULT IS:",datet1)

    return datet1,count1
    
    
     
def Extract_daily():
    count2 = []
    datet2 = []
    mycursor = mydb.cursor()
    query = "SELECT Datetime, DAY_Count FROM log_daily ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results = mycursor.fetchall()
    
    for result in results:
        time = result[0]
        time = datetime.strptime(str(time), "%Y-%m-%d %H:%M:%S")
        datet2.append(time)
    
    for result in results:
        count2.append(result[1])
    
 

    return datet2, count2

def Extract_weekly():

    count3=[]
    datet3=[]

    mycursor = mydb.cursor()
    query = "SELECT Datetime,Week_Count FROM log_weekly ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet3.append(time)
    for result in results:
        count3.append(result[1])
    

    return datet3,count3
     
def Extract_monthly():

    count4=[]
    datet4=[]

    mycursor = mydb.cursor()
    query = "SELECT Datetime,Month_Count FROM log_monthly ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet4.append(time)
    for result in results:
        count4.append(result[1])
    

    return datet4,count4

def Extract_hourly():
    count5=[]
    datet5=[]
    mycursor = mydb.cursor()
    query = "SELECT Datetime,Hour_Count FROM log_hourly_reset ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    # time = result[0]
    # time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
    # result1= result[1]

    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet5.append(time)
    for result in results:
        count5.append(result[1])
    # print("RESULT IS:",datet1)

    return datet5,count5
    
    
     
def Extract_daily():
    count6 = []
    datet6 = []
    mycursor = mydb.cursor()
    query = "SELECT Datetime, DAY_Count FROM log_daily_reset ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results = mycursor.fetchall()
    
    for result in results:
        time = result[0]
        time = datetime.strptime(str(time), "%Y-%m-%d %H:%M:%S")
        datet6.append(time)
    
    for result in results:
        count6.append(result[1])
    
 

    return datet6, count6

def Extract_weekly():

    count7=[]
    datet7=[]

    mycursor = mydb.cursor()
    query = "SELECT Datetime,Week_Count FROM log_weekly_reset ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet7.append(time)
    for result in results:
        count7.append(result[1])
    

    return datet7,count7
     
def Extract_monthly():

    count8=[]
    datet8=[]

    mycursor = mydb.cursor()
    query = "SELECT Datetime,Month_Count FROM log_monthly_reset ORDER BY Datetime DESC LIMIT 30"
    mycursor.execute(query)
    results=mycursor.fetchall()
    # result = tuple(0 if value is None else value for value in result)
    for result in results:
        time = result[0]
        time=datetime.strptime(str(time),"%Y-%m-%d %H:%M:%S")
        datet8.append(time)
    for result in results:
        count8.append(result[1])
    

    return datet8,count8
 
def Extract():


    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="Hashini@123",
        database="coconut"
    )
    
    try:
       
        mycursor = mydb.cursor()
        query = "SELECT Count FROM data ORDER BY Datetime DESC LIMIT 1"
        mycursor.execute(query)
        result = mycursor.fetchone()
    
        if result:
            return result[0]
        else:
            # Handle the case when there's no data in the table
            return 0  # Or any default value you prefer
        

    finally:
        mydb.close()  # Always close connections

    

def Extract1():

    mycursor = mydb.cursor()
    query = "SELECT Cummulative_Count FROM totalizer ORDER BY Datetime DESC LIMIT 1"
    mycursor.execute(query)
    result=mycursor.fetchone()
    result= result[0]
    
    # print(result)

    return result

def DELETE(time):
    mycursor = mydb.cursor()
    clear_query = "DELETE FROM DATA"
    mycursor.execute(clear_query)
    mydb.commit()
    reset(time)


def WeekCount():
    mycursor = mydb.cursor()
    query = "SELECT DAY_Count FROM LOG ORDER BY Datetime DESC LIMIT 7"
    mycursor.execute(query)

    results=mycursor.fetchall()
    value=[result[0] for result in results]

    total= sum(value)

    return total

def MonthCount():
    mycursor = mydb.cursor()
    query = "SELECT Datetime,Week_Count FROM LOG WHERE Datetime >= DATE_ADD(CURDATE(), INTERVAL -1 MONTH) AND Datetime < CURDATE() ORDER BY Datetime DESC;"
    mycursor.execute(query)

    results=mycursor.fetchall()
    value=[result[1] for result in results]

    total= sum(value)

    return total

def current_month():
    mycursor = mydb.cursor()
    query = "SELECT Datetime FROM LOG ORDER BY Datetime DESC LIMIT 1"
    mycursor.execute(query)

    result = mycursor.fetchone()

    if result:
        # Extract month value from the datetime
        datetime_value = result[0]
        month_value = datetime_value.month
        return month_value
    else:
        # Handle the case where no entry is found in the LOG table
        return None
def store_coconut_count(count):
    insert_coconut_count(count) 

def log_resets():
    # schedule.every().hour.do(store_hourly_resets)
    schedule.every().hour.at(":00").do(store_hourly_resets)
    print("log_resets")
    schedule.every().day.at("00:00").do(store_daily_resets)
    # schedule.every().month.do(store_monthly_resets)
    schedule.every(7).days.do(store_weekly_resets)   
    

def log_day_resets():
    
    store_daily_resets()
    store_monthly_resets()

def log_month_resets():
    

    store_monthly_resets()

def Hour_createcsv(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_hourly WHERE Datetime > '{i}' AND Datetime < '{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Hour_Count FROM log_hourly WHERE Datetime>'{i}' AND Datetime < '{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Hour_report.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Hour_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save


def Hour_createcsv_reset(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of hour reset count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_hourly_reset WHERE Datetime > '{i}' AND Datetime < '{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
        
    query1= f"SELECT Hour_Count FROM log_hourly_reset WHERE Datetime>'{i}' AND Datetime < '{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Hour_report_reset.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files"
    file_path = os.path.join(destinationfolder, CSV_file)
    print("happy")

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Hour_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
            
            
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save
        

    # for row in rows:
    #     with open(destination_file_path, 'w', newline='')as file:
    #         csv_writer = csv.writer(file)
    #         csv_writer.writerow(row)
                
    # message = MIMEMultipart()
    # message['From']= sender_email
    # message['To']= reciever_email
    # message['Subject']= subject
    # message.attach(MIMEText(body,'plain'))

    # with open(attatchment_path, 'rb') as file:
    #     attachment = MIMEApplication(file.read(), _subtype='csv')
    #     attachment.add_header('Content-Disposition', f'attachment; filename=report_20240307.csv')
    #     message.attach(attachment)

    #     # Email server configuration (for Gmail)
    # smtp_server = 'smtp.gmail.com'
    # smtp_port = 587
    # smtp_username = 'your_email@gmail.com'
    # smtp_password = 'your_email_password'

    # # Start the SMTP server
    # server = smtplib.SMTP(smtp_server, smtp_port)
    # server.starttls()

    # # Login to the email server
    # server.login(smtp_username, smtp_password)

    # # Send the email
    # server.sendmail(sender_email, reciever_email, message.as_string())

    # # Quit the server
    # server.quit()

def Day_createcsv(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_daily WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT DAY_Count FROM log_daily WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_DAY_report.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Day_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save


def Day_createcsv_reset(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_daily_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT DAY_Count FROM log_daily_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_DAY_report_reset.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reset_Reports"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Day_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save

def Week_createcsv(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_weekly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Week_Count FROM log_weekly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Week_report.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Week_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save


def Week_createcsv_reset(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_weekly_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Week_Count FROM log_weekly_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Week_report_reset.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reset_Reports"
 
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Week_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save

def Month_createcsv(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_monthly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Month_Count FROM log_monthly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Month_report.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Month_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save

def Month_createcsv_reset(time1,time2): 

    date_list=[]
    count_list=[]
    sender_email = ''
    reciever_email=''
    subject = 'Report of daily count'
    body = 'Please find the attatched CSV report'

    current_time = datetime.now().strftime('%Y-%m-%d-%H')
    mycursor = mydb.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_monthly_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Month_Count FROM log_monthly_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    CSV_file = f'{current_time}_Month_report_reset.xlsx'

    destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports/Reset_Reports"
    file_path = os.path.join(destinationfolder, CSV_file)

    # Creating a DataFrame
    df = pd.DataFrame({'Datetime': date_list, 'Month_Count': count_list})

    with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
        
        
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet['1:1']:
            cell.fill = header_fill
            cell.border = header_border

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value is not None]
            if column:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        writer._save

def createcsv(value): 

    if value ==1:

        date_list=[]
        count_list=[]
        sender_email = ''
        reciever_email=''
        subject = 'Report of daily count'
        body = 'Please find the attatched CSV report'

        current_time = datetime.now().strftime('%Y-%m-%d-%H')
        mycursor = mydb.cursor()

        query = f"SELECT Datetime FROM log_daily"
        mycursor.execute(query)

        date = mycursor.fetchall()
        for d in date:
            date_list.append(d[0])
        
        query1= f"SELECT DAY_Count FROM log_daily"
        mycursor.execute(query1)

        count = mycursor.fetchall()
        for c in count:
            count_list.append(c[0])

        CSV_file = f'{current_time}_Complete_DAY_report.xlsx'

        destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
        file_path = os.path.join(destinationfolder, CSV_file)

        # Creating a DataFrame
        df = pd.DataFrame({'Datetime': date_list, 'Day_Count': count_list})

        with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
            
            
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in worksheet['1:1']:
                cell.fill = header_fill
                cell.border = header_border

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column if cell.value is not None]
                if column:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            writer._save

    if value ==2:

        date_list=[]
        count_list=[]
        sender_email = ''
        reciever_email=''
        subject = 'Report of daily count'
        body = 'Please find the attatched CSV report'

        current_time = datetime.now().strftime('%Y-%m-%d-%H')
        mycursor = mydb.cursor()

        query = f"SELECT Datetime FROM log_hourly"
        mycursor.execute(query)

        date = mycursor.fetchall()
        for d in date:
            date_list.append(d[0])
        
        query1= f"SELECT Hour_Count FROM log_hourly"
        mycursor.execute(query1)

        count = mycursor.fetchall()
        for c in count:
            count_list.append(c[0])

        CSV_file = f'{current_time}_Complete_Hour_report.xlsx'

        destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
        file_path = os.path.join(destinationfolder, CSV_file)

        # Creating a DataFrame
        df = pd.DataFrame({'Datetime': date_list, 'Hour_Count': count_list})

        with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
            
            
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in worksheet['1:1']:
                cell.fill = header_fill
                cell.border = header_border

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column if cell.value is not None]
                if column:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            writer._save

    if value ==3:

        date_list=[]
        count_list=[]
        sender_email = ''
        reciever_email=''
        subject = 'Report of daily count'
        body = 'Please find the attatched CSV report'

        current_time = datetime.now().strftime('%Y-%m-%d-%H')
        mycursor = mydb.cursor()

        query = f"SELECT Datetime FROM log_weekly"
        mycursor.execute(query)

        date = mycursor.fetchall()
        for d in date:
            date_list.append(d[0])
        
        query1= f"SELECT Week_Count FROM log_weekly"
        mycursor.execute(query1)

        count = mycursor.fetchall()
        for c in count:
            count_list.append(c[0])

        CSV_file = f'{current_time}_Complete_Week_report.xlsx'

        destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
        file_path = os.path.join(destinationfolder, CSV_file)

        # Creating a DataFrame
        df = pd.DataFrame({'Datetime': date_list, 'Week_Count': count_list})

        with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
            
            
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in worksheet['1:1']:
                cell.fill = header_fill
                cell.border = header_border

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column if cell.value is not None]
                if column:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            writer._save
    
    if value ==4:

        date_list=[]
        count_list=[]
        sender_email = ''
        reciever_email=''
        subject = 'Report of daily count'
        body = 'Please find the attatched CSV report'

        current_time = datetime.now().strftime('%Y-%m-%d-%H')
        mycursor = mydb.cursor()

        query = f"SELECT Datetime FROM log_monthly"
        mycursor.execute(query)

        date = mycursor.fetchall()
        for d in date:
            date_list.append(d[0])
        
        query1= f"SELECT Week_Count FROM log_monthly"
        mycursor.execute(query1)

        count = mycursor.fetchall()
        for c in count:
            count_list.append(c[0])

        CSV_file = f'{current_time}_Complete_Month_report.xlsx'

        destinationfolder= "C:/Hashini/AV_IOT/coconut/files/Reports"
        file_path = os.path.join(destinationfolder, CSV_file)

        # Creating a DataFrame
        df = pd.DataFrame({'Datetime': date_list, 'Month_Count': count_list})

        with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
            
            
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in worksheet['1:1']:
                cell.fill = header_fill
                cell.border = header_border

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column if cell.value is not None]
                if column:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            writer._save


# # def createcsv_custom():

#     current_time = datetime.now().strftime('%Y-%m-%d-%H')
#     mycursor = mydb.cursor()
#     query = "SELECT Datetime,DAY_Count FROM LOG"
#     mycursor.execute(query)

#     rows = mycursor.fetchall()

#     CSV_file = f'{current_time}_Month_report.csv'

#     destinationfolder= "C:/Hashini/AV_IOT/coconut/files/total_count_reports"

#     shutil.os.makedirs(destinationfolder,exist_ok=True)

#     destination_file_path = f'{destinationfolder}/{ CSV_file }'

#     with open(destination_file_path, 'w', newline='')as file:
#             csv_writer = csv.writer(file)
#             header = ['Date_Time','Total Count of Coconut']
#             csv_writer.writerow(header)

#             for row in rows:
#                 formatted_date = datetime.datetime.strptime(str(row[0]),"%Y-%m-%d %H:%M:%S")
#                 formatted_date = formatted_date.strftime('%Y-%m-%d %H:%M:%S')
#                 csv_writer.writerow([formatted_date, row[1]])   

# print(Extract1())

# Hour_createcsv('2024-03-13 14:00:28','2024-03-14 12:00:08')``
mydb_hour = mysql.connector.connect(
        host = "localhost",
        user = "COCONUT",
        passwd = "Hashini@123",
        database = "coconut"
)

mydb_day = mysql.connector.connect(
        host = "localhost",
        user = "COCONUT",
        passwd = "Hashini@123",
        database = "coconut"
)

mydb_week = mysql.connector.connect(
        host = "localhost",
        user = "COCONUT",
        passwd = "Hashini@123",
        database = "coconut"
)

mydb_month = mysql.connector.connect(
        host = "localhost",
        user = "COCONUT",
        passwd = "Hashini@123",
        database = "coconut"
)

def Hour_data(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_hour.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_hourly WHERE Datetime > '{i}' AND Datetime < '{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Hour_Count FROM log_hourly WHERE Datetime>'{i}' AND Datetime < '{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])
    
    return date_list,count_list 

def Hour_data_reset(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_hour.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_hourly_reset WHERE Datetime > '{i}' AND Datetime < '{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Hour_Count FROM log_hourly_reset WHERE Datetime>'{i}' AND Datetime < '{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])
    
    return date_list,count_list   
    
def Day_data(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_day.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_daily WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT DAY_Count FROM log_daily WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])
    
    return date_list,count_list 
def Day_data_reset(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_day.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_daily_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT DAY_Count FROM log_daily_reset WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])
    
    return date_list,count_list 
   

def Week_data(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_week.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_weekly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Week_Count FROM log_weekly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    return date_list,count_list 

    

def Month_data(time1,time2): 

    date_list=[]
    count_list=[]
    
    mycursor = mydb_month.cursor()
    i= time1.strftime("%Y-%m-%d %H:%M:%S")
    j= time2.strftime("%Y-%m-%d %H:%M:%S")

    query = f"SELECT Datetime FROM log_monthly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query)

    date = mycursor.fetchall()
    for d in date:
        date_list.append(d[0])
    
    query1= f"SELECT Month_Count FROM log_monthly WHERE Datetime>'{i}' AND Datetime<'{j}'"
    mycursor.execute(query1)

    count = mycursor.fetchall()
    for c in count:
        count_list.append(c[0])

    return date_list,count_list             
